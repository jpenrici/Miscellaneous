# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from PIL import Image, ImageDraw
from pathlib import Path


def generate_placeholder_logo(path: Path):
    """
    Generates a simple placeholder image using Pillow.
    """
    # Create a 200x200 blue square
    img = Image.new('RGB', (200, 200), color=(79, 129, 189))
    d = ImageDraw.Draw(img)

    # Draw a white border and some text
    d.rectangle([10, 10, 190, 190], outline=(255, 255, 255), width=5)
    d.text((50, 90), "LOGO", fill=(255, 255, 255))

    img.save(path)
    print(f"[*] Placeholder logo generated at: {path}")


def create_sample():
    """
    Generates a XLSX to test the XlsxMapper pipeline.
    Includes: Multi-sheet, cross-sheet formulas, nested merges, and various colors.
    """

    # Paths
    base_path = Path(__file__).parent.parent
    test_dir = base_path / "test_xlsx"
    img_path = test_dir / "logo.png"

    test_dir.mkdir(parents=True, exist_ok=True)

    # Generate the Fake Logo
    generate_placeholder_logo(img_path)

    # Workbook
    wb = openpyxl.Workbook()

    # --- SHEET 1: Financial Summary ---
    ws1 = wb.active
    ws1.title = "Financeiro"

    # Styles
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")  # Blue
    header_font = Font(bold=True, color="FFFFFF")
    border_thick = Side(border_style="thick", color="000000")

    # Merged Header
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "ORÇAMENTO ANUAL 2026"
    ws1["A1"].fill = header_fill
    ws1["A1"].font = header_font
    ws1["A1"].alignment = Alignment(horizontal="center")

    # Data and Formulas
    ws1["A2"], ws1["B2"] = "Categoria", "Quantia"
    ws1["A3"], ws1["B3"] = "Marketing", 5000
    ws1["A4"], ws1["B4"] = "TI", 12000

    # Formula cell
    ws1["A5"] = "TOTAL"
    ws1["B5"] = "=SUM(B3:B4)"
    ws1["B5"].font = Font(bold=True)
    ws1["B5"].border = Border(top=border_thick)

    # --- SHEET 2: Matrix_Layout ---
    ws2 = wb.create_sheet("Matriz")

    # Vertical Merge
    ws2.merge_cells("A2:A10")
    ws2["A2"] = "Texto Vertical"
    ws2["A2"].alignment = Alignment(textRotation=90, vertical="center", horizontal="center")
    ws2["A2"].fill = PatternFill(start_color="E1E1E1", fill_type="solid")  # Grey

    # Horizontal Merge with specific color
    ws2.merge_cells("C3:F3")
    ws2["C3"] = "TÍTULO"
    ws2["C3"].fill = PatternFill(start_color="FFC000", fill_type="solid")  # Orange/Gold
    ws2["C3"].alignment = Alignment(horizontal="center")

    # Cross-Sheet Formula: Reference value from Sheet 1
    ws2["B1"] = "Financeiro (Total):"
    ws2["C1"] = "=Financeiro!B5"

    # Border Stress Test
    red_side = Side(border_style="medium", color="FF0000")
    ws2["D5"].value = "CAIXA VERMELHA"
    ws2["D5"].border = Border(top=red_side, bottom=red_side, left=red_side, right=red_side)

    # --- SHEET 3: Image ---
    ws3 = wb.create_sheet("Imagem")

    ws3["A1"] = "LOGO"
    ws3["A1"].font = Font(bold=True, size=14)

    # Insert Image - Anchor it at column D, row 2
    img = XLImage(img_path)
    ws3.add_image(img, "D2")

    # Save
    output_path = test_dir / "sample.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[+] Sample created at: {output_path}")


if __name__ == "__main__":
    create_sample()
