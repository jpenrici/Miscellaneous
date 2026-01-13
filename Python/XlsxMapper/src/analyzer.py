# -*- coding: utf-8 -*-

import json
import openpyxl
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import List, Optional, Dict, Any


@dataclass
class CellMetadata:
    """
    Data structure representing exhaustive metadata for a single Excel cell.
    Used for both JSON persistence and as input for various exporters.
    """
    coordinate: str
    row: int
    col: int
    value: Any
    is_merged: bool
    merge_range: Optional[str] = None
    fill_color: Optional[str] = None   # Stored as HEX (e.g., 'FF0000')
    font_bold: bool = False
    alignment: str = "left"
    formula: Optional[str] = None
    # Dictionary to store border styles: {'top': {'style': 'thin', 'color': '000000'}, ...}
    borders: Dict[str, Any] = field(default_factory=dict)


class XlsxAnalyzer:
    """
    Core engine responsible for parsing XLSX files and extracting 
    content and formatting metadata.
    """

    def __init__(self, file_path: Path):
        self.file_path = file_path
        # Load with data_only=False to ensure formulas are captured as strings
        self.workbook = openpyxl.load_workbook(file_path, data_only=False)

    def _get_hex_color(self, color_obj: Any) -> Optional[str]:
        """
        Converts openpyxl color objects (ARGB) to standard HEX strings.
        Fixes the 'black background' issue by checking for invalid/default colors.
        """
        if color_obj and color_obj.type == 'rgb':
            rgb = str(color_obj.rgb)
            # If color is '00000000' or empty, it means 'No Fill' or 'Automatic'
            if rgb in ["00000000", "000000", "None"]:
                return None

            return rgb[2:] if len(rgb) == 8 else rgb
        return None

    def _get_borders(self, cell: Any) -> Dict[str, Any]:
        """
        Maps the border style and color for each side of the cell.
        """
        border_map = {}
        for side_name in ['top', 'bottom', 'left', 'right']:
            side = getattr(cell.border, side_name)
            if side and side.style:
                border_map[side_name] = {
                    "style": side.style,
                    "color": self._get_hex_color(side.color) or "000000"
                }
        return border_map

    def get_sheet_dimensions(self, sheet_name: str) -> Dict[str, Any]:
        """Captures column widths and row heights for visual fidelity."""
        ws = self.workbook[sheet_name]
        dims = {"cols": {}, "rows": {}}

        for col_letter, col_def in ws.column_dimensions.items():
            if col_def.width:
                dims["cols"][col_letter] = col_def.width

        for row_idx, row_def in ws.row_dimensions.items():
            if row_def.height:
                dims["rows"][row_idx] = row_def.height
        return dims

    def get_cell_details(self, sheet_name: str) -> List[CellMetadata]:
        """
        Iterates through the worksheet and compiles a list of CellMetadata objects.
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} not found in workbook.")

        ws = self.workbook[sheet_name]
        mapped_data: List[CellMetadata] = []

        # Get all merged ranges to check against during iteration
        merged_ranges = ws.merged_cells.ranges

        # Safety limits to prevent processing massive empty areas
        max_r = min(ws.max_row, 200)
        max_c = min(ws.max_column, 26)

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)

                # Check for formula: value starts with '=' in data_only=False mode
                raw_value = cell.value
                formula = raw_value if isinstance(raw_value, str) and raw_value.startswith('=') else None

                # Identify if cell belongs to a merged range
                m_range_coord = None
                is_merged = False
                for m_range in merged_ranges:
                    if cell.coordinate in m_range:
                        is_merged = True
                        m_range_coord = m_range.coord
                        break

                meta = CellMetadata(
                    coordinate=cell.coordinate,
                    row=r,
                    col=c,
                    value=raw_value if not formula else None,
                    formula=formula,
                    is_merged=is_merged,
                    merge_range=m_range_coord,
                    font_bold=getattr(cell.font, "bold", False),
                    fill_color=self._get_hex_color(cell.fill.start_color),
                    alignment=getattr(cell.alignment, "horizontal", "left") or "left",
                    borders=self._get_borders(cell)
                )
                mapped_data.append(meta)

        return mapped_data

    def export_config_json(self, sheet_name: str, output_path: Path) -> None:
        """
        Serializes the extracted metadata to a JSON file.
        """
        data = self.get_cell_details(sheet_name)
        # Convert list of dataclasses to a list of dictionaries
        serializable_data = [asdict(item) for item in data]

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(serializable_data, f, indent=4)
