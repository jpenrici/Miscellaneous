# -*- coding: utf-8 -*-

import openpyxl

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pathlib import Path


def create_test_file(file_path: str) -> None:
    """
    Creates a complex xlsx file with various formats, colors, and borders
    to challenge the transpiler.
    """
    wb = openpyxl.Workbook()

    # 1. Sheet: Styling & Borders
    ws1 = wb.active
    ws1.title = "Formatted_Data"

    # Define some reusable styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Define borders (Thin and Thick)
    thin_side = Side(border_style="thin", color="000000")
    thick_side = Side(border_style="thick", color="FF0000")
    standard_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    alert_border = Border(top=thick_side, left=thick_side, right=thick_side, bottom=thick_side)

    # Header with Merge and Style
    ws1.merge_cells("A1:D1")
    header_cell = ws1["A1"]
    header_cell.value = "Inventory Report 2024"
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.font = header_font
    header_cell.fill = header_fill
    header_cell.border = standard_border

    # Data Rows
    ws1["A2"] = "Code"
    ws1["B2"] = "Product"
    ws1["C2"] = "Quantity"
    ws1["D2"] = "Alert"

    # Applying a red alert style to a specific cell
    ws1["D3"] = "LOW STOCK"
    ws1["D3"].font = Font(color="FF0000", bold=True)
    ws1["D3"].border = alert_border
    ws1["D3"].alignment = Alignment(horizontal="right")

    # 2. Sheet: Merged Complexity
    ws2 = wb.create_sheet("Matrix_Layout")

    # Horizontal Merge
    ws2.merge_cells("B2:E2")
    ws2["B2"] = "Horizontal Group"
    ws2["B2"].alignment = Alignment(horizontal="center")
    ws2["B2"].fill = PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")

    # Vertical Merge
    ws2.merge_cells("A2:A10")
    ws2["A2"] = "Side Sidebar"
    ws2["A2"].alignment = Alignment(text_rotation=90, vertical="center", horizontal="center")

    # Final save
    file_path = OUTPUT_PATH/file_path
    wb.save(file_path)
    print(f"Test file created: {file_path}")


if __name__ == "__main__":

    ROOT: Path = Path(__file__).resolve().parent
    OUTPUT_PATH = ROOT / "../test_xlsx"

    OUTPUT_PATH.mkdir(parents=True, exist_ok=True)

    create_test_file("test_spreadsheet.xlsx")
