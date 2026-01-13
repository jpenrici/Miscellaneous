# -*- coding: utf-8 -*-

import openpyxl
import os
import sys

from pathlib import Path
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from typing import List


def get_cell_value(sheet: Worksheet, row: int, col: int) -> str:
    """
    Checks if a cell is merged and returns the primary value.

    Args:
        sheet: The openpyxl worksheet object.
        row: Row index.
        col: Column index.

    Returns:
        String representation of the cell value.
    """
    cell: Cell = sheet.cell(row=row, column=col)

    # Check if the cell is part of any merged range
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Always fetch the value from the top-left cell of the range
            merged_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return str(merged_value) if merged_value is not None else ""

    return str(cell.value) if cell.value is not None else ""


def generate_ascii_table(sheet_data: List[List[str]]) -> str:
    """
    Transforms a 2D list of strings into a formatted ASCII table.
    """
    if not sheet_data:
        return ""

    # Calculate the maximum width for each column
    col_widths: List[int] = [
        max(len(row[i]) for row in sheet_data)
        for i in range(len(sheet_data[0]))
    ]

    # Create the horizontal border line (+---+-------+)
    separator: str = "+" + "+".join("-" * (w + 2) for w in col_widths) + "+"

    lines: List[str] = [separator]
    for row in sheet_data:
        # Create the data line (| val | val |)
        content_line: str = "|" + "|".join(
            f" {row[i].ljust(col_widths[i])} " for i in range(len(row))
        ) + "|"
        lines.append(content_line)
        lines.append(separator)

    return "\n".join(lines)


def process_workflow(file_path: str) -> None:
    """
    Main workflow to read xlsx and save txt tables.
    """
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.")
        return

    try:
        # data_only=True helps getting values instead of formulas
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        print(f"File loaded. Sheets found: {len(workbook.sheetnames)}")

        for name in workbook.sheetnames:
            ws: Worksheet = workbook[name]

            # Limit dimensions to prevent memory issues with massive files
            max_r: int = min(ws.max_row, 200)
            max_c: int = min(ws.max_column, 26)

            print(f"Processing sheet '{name}' ({max_r}x{max_c})...")

            # Extract data into a 2D List
            table_matrix: List[List[str]] = []
            for r in range(1, max_r + 1):
                row_content: List[str] = []
                for c in range(1, max_c + 1):
                    row_content.append(get_cell_value(ws, r, c))
                table_matrix.append(row_content)

            # Convert to ASCII and Save
            ascii_art: str = generate_ascii_table(table_matrix)
            output_path: str = f"table_{name}.txt"

            with open(output_path, "w", encoding="utf-8") as f:
                f.write(ascii_art)

            print(f"Success: {output_path} created.")

    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        process_workflow(sys.argv[1])
    else:
        script_name = Path(sys.argv[0]).name if sys.argv[0] else "script.py"
        print(f"Usage: python {script_name} <file_path.xlsx>")
