# -*- coding: utf-8 -*-

from pathlib import Path
from typing import Dict, Any


class PythonScriptExporter:
    """Generates a master Python script to mirror the original XLSX workbook."""

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir

    def generate_full_workbook(self, workbook_data: Dict[str, Dict[str, Any]]) -> None:
        """
        Generates a master script to rebuild the workbook.
        Expected structure: { "SheetName": {"cells": [...], "dims": {...}, "assets": [...] } }
        """
        script = [
            "# -*- coding: utf-8 -*-",
            "import openpyxl",
            "from openpyxl.styles import Alignment, Font, PatternFill, Border, Side",
            "from openpyxl.drawing.image import Image as XLImage\n",
            "def rebuild():",
            "    print('[*] Starting Workbook reconstruction...')",
            "    wb = openpyxl.Workbook()",
            "    # Remove default sheet",
            "    if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])\n"
        ]

        for sheet_name, content in workbook_data.items():
            script.append(f"    # --- Reconstructing Sheet: {sheet_name} ---")
            script.append(f"    ws = wb.create_sheet('{sheet_name}')")

            # Apply Dimensions
            dims = content.get("dims", {})
            for col, width in dims.get("cols", {}).items():
                script.append(f"    ws.column_dimensions['{col}'].width = {width}")
            for row, height in dims.get("rows", {}).items():
                script.append(f"    ws.row_dimensions[{row}].height = {height}")

            # Reconstruct Cells (Data & Styles)
            processed_merges = set()
            cells = content.get("cells", [])

            for c in cells:
                # Support both Dict (from JSON) and Dataclass
                data = c if isinstance(c, dict) else c.__dict__
                coord = data['coordinate']

                # Formulas and Values
                if data.get('formula'):
                    script.append(f"    ws['{coord}'].value = '{data['formula']}'")
                elif data.get('value') is not None:
                    val = data['value']
                    formatted_val = f'"""{val}"""' if isinstance(val, str) else val
                    script.append(f"    ws['{coord}'].value = {formatted_val}")

                # Alignment & Rotation
                align_parts = []
                if data.get('alignment') != 'left':
                    align_parts.append(f"horizontal='{data['alignment']}'")
                if data.get('vertical_align') and data.get('vertical_align') != 'bottom':
                    align_parts.append(f"vertical='{data['vertical_align']}'")
                if data.get('text_rotation', 0) != 0:
                    align_parts.append(f"text_rotation={data['text_rotation']}")

                if align_parts:
                    script.append(f"    ws['{coord}'].alignment = Alignment({', '.join(align_parts)})")

                # Borders
                if data.get('borders'):
                    b_parts = [f"{s}=Side(border_style='{d['style']}', color='{d['color']}')"
                               for s, d in data['borders'].items()]
                    script.append(f"    ws['{coord}'].border = Border({', '.join(b_parts)})")

                # Fill & Font
                if data.get('fill_color'):
                    script.append(
                        f"    ws['{coord}'].fill = PatternFill(start_color='{data['fill_color']}', fill_type='solid')")
                if data.get('font_bold'):
                    script.append(f"    ws['{coord}'].font = Font(bold=True)")

                # Merge Logic
                m_range = data.get('merge_range')
                if data.get('is_merged') and m_range not in processed_merges:
                    script.append(f"    ws.merge_cells('{m_range}')")
                    processed_merges.add(m_range)

            # 3. Reconstruct Assets (Images)
            assets = content.get("assets", [])
            if assets:
                script.append(f"    # Assets for {sheet_name}")
                for img in assets:
                    # We assume 'images/' is a subfolder where the script is run
                    script.append("    try:")
                    script.append(f"        img_obj = XLImage('images/{img['filename']}')")
                    script.append(f"        img_obj.width, img_obj.height = {img['width']}, {img['height']}")
                    script.append(f"        ws.add_image(img_obj, '{img['anchor']}')")
                    script.append(
                        f"    except Exception as e: print(f' [!] Error loading image {img['filename']}: {{e}}')")

            script.append("")  # Spacer between sheets

        script.extend([
            "    output_filename = 'full_reconstructed_file.xlsx'",
            "    wb.save(output_filename)",
            "    print(f\"[+] Success! '{output_filename}' created.\")",
            "\nif __name__ == '__main__':",
            "    rebuild()"
        ])

        output_file = self.output_dir / "full_reconstruct.py"
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(script))


if __name__ == "__main__":
    print("Module to convert cell metadata or JSON, generated by the Analysis Module,"
          " into a Python script using Openpyxl."
          "\nUse:\n\tfrom exporter_py import PythonScriptExporter")
