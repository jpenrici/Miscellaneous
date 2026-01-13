# -*- coding: utf-8 -*-

import openpyxl
import sys

from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from typing import List


def translate_to_python(file_path: str) -> None:
    """Reads an xlsx and generates a Python script to recreate it."""
    wb = openpyxl.load_workbook(file_path)

    script_content: List[str] = [
        "import openpyxl",
        "from openpyxl.styles import Alignment, PatternFill, Font\n",
        "def rebuild_spreadsheet():",
        "    wb = openpyxl.Workbook()",
        "    # Remove default sheet",
        "    default_std = wb.active",
        "    wb.remove(default_std)\n"
    ]

    for sheet_name in wb.sheetnames:
        ws: Worksheet = wb[sheet_name]
        script_content.append(f"    # --- Processing Sheet: {sheet_name} ---")
        script_content.append(f"    ws = wb.create_sheet(title='{sheet_name}')")

        # Handle Merged Cells first
        for merged_range in ws.merged_cells.ranges:
            script_content.append(f"    ws.merge_cells('{merged_range.coord}')")

        # Handle Cell Values and basic formatting
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Escape strings to avoid syntax errors
                    val = f"'{cell.value}'" if isinstance(cell.value, str) else cell.value
                    script_content.append(f"    ws['{cell.coordinate}'].value = {val}")

                # Copy basic alignment if it exists
                if cell.alignment and cell.alignment.horizontal:
                    script_content.append(
                        f"    ws['{cell.coordinate}'].alignment = "
                        f"Alignment(horizontal='{cell.alignment.horizontal}')"
                    )

        script_content.append("")  # Spacer

    script_content.extend([
        "    wb.save('reconstructed_file.xlsx')",
        "    print('File reconstructed successfully!')\n",
        "if __name__ == '__main__':",
        "    rebuild_spreadsheet()"
    ])

    # Save the generated script
    with open("reconstruct_table.py", "w", encoding="utf-8") as f:
        f.write("\n".join(script_content))

    print("Done! 'reconstruct_table.py' has been generated.")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        translate_to_python(sys.argv[1])
    else:
        script_name = Path(sys.argv[0]).name if sys.argv[0] else "script.py"
        print(f"Usage: python {script_name} <file_path.xlsx>")
